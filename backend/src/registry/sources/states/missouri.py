from __future__ import annotations

import asyncio
import hashlib
import io
import re
import zipfile
from collections import defaultdict
from collections.abc import AsyncIterator
from datetime import date, datetime
from typing import Any

import httpx
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from registry.sources.base import SourceConnector


def _compact_text(value: str | None) -> str | None:
    if value is None:
        return None
    text = " ".join(value.split())
    return text or None


def _to_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = _compact_text(str(value))
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _parse_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _compact_text(str(value))
    if not text:
        return None
    text = text.replace("/", "-")
    if text.endswith(" 00:00:00"):
        text = text[:10]
    for fmt in ("%Y-%m-%d", "%m-%d-%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, (int, float, bool)):
        return value
    return _compact_text(str(value))


def _normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    return {key: _cell_value(value) for key, value in row.items() if key}


def _row_key(row: dict[str, Any]) -> str:
    return "|".join(
        [
            _compact_text(str(row.get("Name") or "")) or "",
            _compact_text(str(row.get("Address") or "")) or "",
            _compact_text(str(row.get("City") or "")) or "",
            _compact_text(str(row.get("St") or "")) or "",
            _compact_text(str(row.get("Zip") or "")) or "",
            _compact_text(str(row.get("County") or "")) or "",
            _compact_text(str(row.get("Date of Birth") or "")) or "",
        ]
    )


def _reduced_row_key(row: dict[str, Any]) -> str:
    parts = [
        _compact_text(str(row.get("Address") or "")) or "",
        _compact_text(str(row.get("City") or "")) or "",
        _compact_text(str(row.get("St") or "")) or "",
        _compact_text(str(row.get("Zip") or "")) or "",
        _compact_text(str(row.get("County") or "")) or "",
        _compact_text(str(row.get("Date of Birth") or "")) or "",
    ]
    return "|".join(parts)


def _external_id(key: str) -> str:
    return f"mo:{hashlib.sha1(key.encode('utf-8')).hexdigest()}"


def _parse_workbook_rows(data: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    worksheet = workbook.active

    headers: list[str] | None = None
    rows: list[dict[str, Any]] = []
    for raw_row in worksheet.iter_rows(values_only=True):
        values = [_cell_value(cell) for cell in raw_row]
        if headers is None:
            if values and _compact_text(str(values[0] or "")) == "Name":
                headers = [_compact_text(str(value or "")) or "" for value in values]
            continue
        if not any(value not in (None, "") for value in values):
            continue
        row = {headers[index]: values[index] for index in range(min(len(headers), len(values))) if headers[index]}
        rows.append(_normalize_row(row))
    return rows


def _parse_vehicle_rows(html: str) -> list[dict[str, Any]]:
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table")
    if table is None:
        return []

    headers: list[str] | None = None
    rows: list[dict[str, Any]] = []
    for tr in table.find_all("tr"):
        cells = [_cell_value(cell.get_text(" ", strip=True)) for cell in tr.find_all(["th", "td"])]
        if headers is None:
            if cells and _compact_text(str(cells[0] or "")) == "Name":
                headers = [_compact_text(str(value or "")) or "" for value in cells]
            continue
        if not any(value not in (None, "") for value in cells):
            continue
        row = {headers[index]: cells[index] for index in range(min(len(headers), len(cells))) if headers[index]}
        rows.append(_normalize_row(row))
    return rows


class MissouriRegistryConnector(SourceConnector):
    name = "missouri"
    state = "MO"
    source_url = "https://www.mshp.dps.missouri.gov/MSHPWeb/PatrolDivisions/CRID/SOR/msor.zip"
    page_size_cap = 500
    _cached_records: list[dict[str, Any]] | None = None

    async def _download_archive(self) -> bytes:
        async with httpx.AsyncClient(timeout=httpx.Timeout(60.0, connect=20.0), verify=False) as client:
            response = await client.get(self.source_url, headers={"User-Agent": "Mozilla/5.0"})
            response.raise_for_status()
            return response.content

    async def _records(self) -> list[dict[str, Any]]:
        if self._cached_records is not None:
            return self._cached_records

        archive_bytes = await self._download_archive()
        with zipfile.ZipFile(io.BytesIO(archive_bytes)) as archive:
            master_rows = _parse_workbook_rows(archive.read("msor.xlsx"))
            offense_rows = _parse_workbook_rows(archive.read("msor_offense.xlsx"))
            alias_rows = _parse_workbook_rows(archive.read("msor_alias.xlsx"))
            vehicle_rows = _parse_vehicle_rows(archive.read("msor_veh.xls").decode("windows-1252", errors="replace"))

        grouped: dict[str, dict[str, Any]] = defaultdict(
            lambda: {
                "master_rows": [],
                "offense_rows": [],
                "vehicle_rows": [],
            }
        )
        alias_index: dict[str, list[dict[str, Any]]] = defaultdict(list)

        for row in master_rows:
            grouped[_row_key(row)]["master_rows"].append(row)
        for row in offense_rows:
            grouped[_row_key(row)]["offense_rows"].append(row)
        for row in vehicle_rows:
            grouped[_row_key(row)]["vehicle_rows"].append(row)
        for row in alias_rows:
            alias_index[_reduced_row_key(row)].append(row)

        records: list[dict[str, Any]] = []
        for key, buckets in grouped.items():
            primary = (
                buckets["master_rows"][0]
                if buckets["master_rows"]
                else buckets["offense_rows"][0]
                if buckets["offense_rows"]
                else buckets["vehicle_rows"][0]
            )
            if not primary:
                continue

            alias_rows_for_person = alias_index.get(_reduced_row_key(primary), [])

            records.append(
                {
                    "person_key": key,
                    "primary": primary,
                    "master_rows": buckets["master_rows"],
                    "offense_rows": buckets["offense_rows"],
                    "alias_rows": alias_rows_for_person,
                    "vehicle_rows": buckets["vehicle_rows"],
                    "source_url": self.source_url,
                }
            )

        self._cached_records = records
        return records

    async def fetch(self, *, limit: int | None = None) -> list[dict[str, Any]]:
        records: list[dict[str, Any]] = []
        async for batch, _cursor in self.fetch_batches(limit=limit):
            records.extend(batch)
        return records

    async def fetch_batches(
        self,
        *,
        limit: int | None = None,
        batch_size: int | None = None,
        cursor: str | None = None,
    ) -> AsyncIterator[tuple[list[dict[str, Any]], str | None]]:
        records = await self._records()
        start = int(cursor) if cursor else 0
        if start < 0:
            start = 0

        selected = records[start:]
        if limit is not None:
            selected = selected[:limit]
        if not selected:
            return

        batch_limit = min(batch_size or self.page_size_cap, self.page_size_cap)
        for offset in range(0, len(selected), batch_limit):
            batch = selected[offset : offset + batch_limit]
            next_index = start + offset + len(batch)
            next_cursor = str(next_index) if next_index < start + len(selected) else None
            yield batch, next_cursor

    def parse(self, raw_payloads: list[dict[str, Any]]) -> list[dict[str, Any]]:
        return raw_payloads

    def normalize(self, parsed_records: list[dict[str, Any]]) -> list[dict[str, Any]]:
        normalized: list[dict[str, Any]] = []
        for record in parsed_records:
            primary = record.get("primary") or {}
            person_key = record.get("person_key") or _row_key(primary)
            external_id = _external_id(person_key)

            name = _compact_text(str(primary.get("Name") or "")) or external_id
            address = {
                "line1": primary.get("Address") or None,
                "line2": None,
                "city": primary.get("City") or None,
                "state": primary.get("St") or None,
                "postal_code": primary.get("Zip") or None,
                "county": primary.get("County") or None,
                "address_precision": "registry",
            }
            addresses = [address] if any(value for key, value in address.items() if key != "address_precision") else []

            offense_rows = record.get("offense_rows") or record.get("master_rows") or []
            offenses = []
            for offense_row in offense_rows:
                offenses.append(
                    {
                        "offense_name": offense_row.get("Offense") or "Missouri registry offense",
                        "count": _to_int(offense_row.get("Count")),
                        "compliant": offense_row.get("Compliant") or None,
                        "tier": offense_row.get("Tier") or None,
                        "victim_gender": offense_row.get("Victim Gender") or None,
                        "victim_age": offense_row.get("Victim Age") or None,
                        "victim_max_age": offense_row.get("Victim Max Age") or None,
                        "offense_city": offense_row.get("Offense City") or None,
                        "offense_state": offense_row.get("Offense State") or None,
                        "offense_date": _parse_date(offense_row.get("Offense Date")),
                        "conviction_date": _parse_date(offense_row.get("Conviction Date")),
                        "confinement_release_date": _parse_date(offense_row.get("Confinement Release Date")),
                        "probation_parole_release_date": _parse_date(offense_row.get("Probation/Parole Release Date")),
                        "offender_age_at_time_of_offense": _to_int(offense_row.get("Offender Age at Time of Offense")),
                    }
                )

            aliases = []
            seen_aliases: set[str] = set()
            for alias_row in record.get("alias_rows") or []:
                alias_name = _compact_text(str(alias_row.get("Name") or ""))
                if not alias_name or alias_name == name or alias_name in seen_aliases:
                    continue
                seen_aliases.add(alias_name)
                aliases.append(alias_name)

            vehicles = []
            for vehicle_row in record.get("vehicle_rows") or []:
                vehicles.append(
                    {
                        "vehicle_make": vehicle_row.get("Vehicle Make") or None,
                        "vehicle_model": vehicle_row.get("Vehicle Model") or None,
                        "vehicle_color_code": vehicle_row.get("Vehicle Color Code") or None,
                        "vehicle_color": vehicle_row.get("Vehicle Color") or None,
                        "license_year": _to_int(vehicle_row.get("License Year")),
                        "license_plate": vehicle_row.get("License") or None,
                        "license_state": vehicle_row.get("License State") or None,
                        "vehicle_owner": vehicle_row.get("Vehicle Owner") or None,
                    }
                )

            tier = primary.get("Tier") or None
            birthday = _parse_date(primary.get("Date of Birth"))

            normalized.append(
                {
                    "external_id": external_id,
                    "full_name": name,
                    "date_of_birth": birthday,
                    "risk_level": f"Tier {tier}" if tier else None,
                    "demographics": {
                        "county": primary.get("County") or None,
                        "city": primary.get("City") or None,
                        "state": primary.get("St") or None,
                        "compliant": primary.get("Compliant") or None,
                        "tier": _to_int(tier),
                        "offense_count": len(offenses),
                        "alias_count": len(aliases),
                        "vehicle_count": len(vehicles),
                        "vehicles": vehicles,
                    },
                    "aliases": aliases,
                    "addresses": addresses,
                    "offenses": offenses,
                    "photos": [],
                    "source_url": self.source_url,
                    "raw_payload": {
                        "primary": primary,
                        "offenses": offenses,
                        "aliases": aliases,
                        "vehicles": vehicles,
                    },
                }
            )
        return normalized
